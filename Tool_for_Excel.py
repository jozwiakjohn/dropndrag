#!/usr/bin/env python

# 2014 Nov 18, John Jozwiak for Apple RF WETA (Tiberiu and Mohit and Prasanna)

import sys
import math
# https://automatetheboringstuff.com/chapter12/
sys.path.append('jdcal-1.0')
sys.path.append('et_xmlfile-1.0.1')
sys.path.append('openpyxl-2.3.0')
sys.path.append('h5py-2.4.0')
import jdcal
import et_xmlfile
import openpyxl
# http://docs.h5py.org/en/latest/quick.html
import os
import re
import glob
import time
import copy
import getpass
import json  #  to pretty-print
import pprint
import traceback
import sqlite3
import dbm
from   collections import namedtuple
import file_dependence
import versioning
import powertable_workflow
###################################################################################

Edit = namedtuple( 'Edit' , [ 'doc' , 'sheet' , 'cellname' ] )

###################################################################################

def  now_as_a_string():
  return  time.strftime( '%Y.%m.%d.%H.%M.%S.%Z' )

###################################################################################

def  is_an_int( s ):
  try:
    i = int( s )
    return True
  except:
    return False

###################################################################################

def  is_a_float( s ):
  try:
    f = float( s )
    return True
  except:
    return False

###################################################################################

def  is_a_Number( v ):
  try:
    w = int(v)
  except:              #  wasn't a string of an int, or an int.
    try:
      w = float(v)
    except:            #  wasn't a string of a float, or a float.
      return False
  return True

###################################################################################

def  number( string ):  #  i.e., "4" -> int("4") , "21.4" -> float("21.4")
  if   is_an_int  ( string ):
    return   int  ( string )
  elif is_a_float ( string ):
    return  float ( string )
  else:
    return  None
  
###################################################################################

def  is_a_String( s ):
  return ( (None != s) and ('' != s) and (not is_a_Number(s)) )

###################################################################################

def  cell_type( v ):
  if   is_a_Number ( v ) : return 'number'
  elif is_a_String ( v ) : return 'string'
  else                   : return None

###################################################################################

docs_root = os.path.join( os.path.expanduser('~') , 'Documents' , 'eggshell_document_databases' )

def  docs_root_get():
  global docs_root
  if not os.path.isdir ( docs_root ):
    os.system( 'mkdir -p -m 770 ' + docs_root )
  return docs_root

def  docs_root_set( v ):
  global docs_root
  if not os.path.isdir ( v ):
    os.system( 'mkdir -p -m 770 ' + v )
  docs_root = v

###################################################################################

def  debug_printable_copy_of_dict( d ):
  s = {}
  for x in d:
    if   'password' == x:
      s[x] = '?'
    else:
      s[x] = d[x]
  return s

###################################################################################

def  sanitized( args ):
  return debug_printable_copy_of_dict( args )

###################################################################################

def  purge_item_from_map_of_T_to_list_of_T( item , m ):  #  returns a new dict without item anywhere.
  answer = {}
  for x in m:
    if (x == item):
      continue  #  elide item as a key in the answer dict.
    t = []
    for s in m[x]:
      if (s == item):
        continue  #  elide item in any image-list of m.
      t.append(s)
    if len(t) > 0:
      answer[x] = t
  return answer

###################################################################################

def  coords_to_excel_cell_name( c_int , r_int ):

  ###

  def  coords_to_excel_cell_name_pair( c_int , r_int ):  #  (0,0) is A1:  such are Excel coordinates.

    if (c_int < 0) or (r_int < 0) :
      return [ ( 'Nonexisting Cell at Impossible (column,row) = (%d,%d)' % (c_int,r_int) ) , None ]

    f          = float( c_int )
    c_radix_26 = [
                   int( math.floor( f / float( 26*26 ) )) ,
                   int( math.floor( f / 26.0           )) ,
                   c_int % 26
                 ]

    l_radix_26 = [
                   '' if c_radix_26[0] <= 0 else chr( ord('A') + c_radix_26[0] - 1 ) ,
                   '' if c_radix_26[1] <= 0 else chr( ord('A') + c_radix_26[1] - 1 ) ,
                                                 chr( ord('A') + c_radix_26[2]     )
                 ]

    c_string   = str(l_radix_26[0]) + str(l_radix_26[1]) + str(l_radix_26[2])

    #  Excel indexes (verb) cells starting at 1, but our coords start at 0.

    return ( c_string , 1 + r_int )

  ###

  ln = coords_to_excel_cell_name_pair( c_int , r_int )
  return str(ln[0]) + str(ln[1])

###################################################################################

def  excel_cell_name_to_coords( n ):

  c = re.sub( '[0-9]*' , '' , n )                           #  e.g., 'AB10'     ->  'AB'
  c = [ c[i]                   for i in xrange(len(c)) ]    #  e.g., 'AB'       ->  ['A','B']
  c = [ (ord(c[i]) - ord('A')) for i in xrange(len(c)) ]    #  e.g., ['A','B']  ->  [ 0 , 1 ]

  c = [ (c[i] + (1 if (i < len(c) - 1)  else 0))  for i in xrange(len(c)) ]    #  e.g., [ 1 , 1 ]  ->  [ 1 , 2 ]
  k = [ (26**(len(c)-1-i))*c[i]                   for i in xrange(len(c)) ]
  kk = reduce( lambda x,y:x+y , k )

  r = re.sub( '[A-Z]*' , '' , n )
  r = int(r) - 1

  return ( kk , r )

###################################################################################
if False:  #  reality check...
  for   c in range(80):
    for r in range( 1):
      print c,r,coords_to_excel_cell_name( c,r )
  book  = openpyxl.load_workbook ( 'examples/J82_V6fake.xlsx' )
  print book.sheetnames
  print '# NO LONGER VALID to go further AFTER openpyxl.' ; sys.exit(0)
  sheet = book.sheet_by_index(0)
  print 'sheet.cell(0,0).value = ', sheet.cell(0,0).value
  print
  print 'A1' ,excel_cell_name_to_coords( 'A1'  )
  print 'AA1',excel_cell_name_to_coords( 'AA1' )
  print 'AB1',excel_cell_name_to_coords( 'AB1' )
  print 'BA1',excel_cell_name_to_coords( 'BA1' )
  print 'CA1',excel_cell_name_to_coords( 'CA1' )
  sys.exit(0)

  print '\nHere test coords_to_excel_cell_name...'
  examples = []
  import random

  for   c in range(30):
    r = random.randint(0,9)
    name =          coords_to_excel_cell_name( c,r )
    print  '( %3i , %3i ) <-> %s' % (c,r,name)
    examples.append( name )

  print '\nHere test excel_cell_name_to_coords...'
  for e in examples:
    print ( '%4s  <-> ' % e ) , excel_cell_name_to_coords( e )
###################################################################################

def  generate_column_headers( ncols ):
  h = {}
  for i in xrange(ncols):
    f    = float(i)
    h[i] = [
             int( math.floor( f / float( 26*26 ) )) ,
             int( math.floor( f / 26.0           )) ,
             i % 26
           ]

  labels = {}
  for i in h:

    x = ''
    if h[i][0] > 0:
      x = chr( ord('A') + h[i][0] - 1 )

    y = ''
    if h[i][1] > 0:
      y = chr( ord('A') + h[i][1] - 1 )

    z =   chr( ord('A') + h[i][2] )
    labels[i+1] = x + y + z

  labels[0] = ''

  return labels  #  a map of integer to Excel style column header label

###################################################################################

format_python , format_text = range(2)  #  an Enum as in C, sort of.

output_format = format_text

###################################################################################

def  describe_excel( path , to_stdout = True ):

  try:
    book   = openpyxl.load_workbook( path )
    print '# DEBUG sheet names for',path,'are',book.sheetnames
  except:
    print '# Error : the path %s was not an xslx Excel document.' % path
    print traceback.print_exc(),'\n'
    return

# if  not to_stdout:

#   result = [ path , [] ]
#   for sheet in sheets:
#     result[1].append( [sheet.name , sheet.ncols , sheet.nrows ] )

#   return result

# if  format_python == output_format:

#   wb = {}
#   wb['name'  ] = path
#   wb['sheets'] = sheets
#   print wb
# else:
#   print  'The workbook ' , '"' + path + '"' , ' has ' , len(sheets) , ' individual spreadsheets.'
#   for sheet in sheets:
#     print '  %20s has %d columns and %d rows' % ('"' + sheet.name + '"',sheet.ncols,sheet.nrows)

###################################################################################

def  map_of_sheet_names_to_sheet_dimensions( d ):  #  returns a map of sheetname s to (# cols,# rows) for s.

  def  dimensions_of_sheet( d , sheet ):

    cells = available_cells_in_egg_sheet( d , sheet , justfilesnotpaths = True )
    cells = [ excel_cell_name_to_coords( n ) for n in cells ]

    columns = set()
    rows    = set()
    for (c,r) in cells:
      columns.add(c)
      rows   .add(r)
    columns = max(columns)
    rows    = max(rows   )

    return    [ columns , rows ]

  #######

  if '/' in d: d = d.split('/')[-1]  #  just the root folder for the doc, not the full path.

  if not os.path.exists( os.path.join( docs_root_get() , d ) ):
    print '# Error : do describe eggshell(',d,') but that doc does not exist.'
    return None

  sheets = available_sheets_in_doc( d , justrootsnotpaths = True )
  dims = {}
  for s in sheets:

    dims[s] = dimensions_of_sheet( d , s )

  return dims

###################################################################################

def  evaluate_cell_formula( a , row , col ):
  if is_a_Number( a[row,col] ):
    pass
  else:
    print '# do something complex'

###################################################################################

def  write_in_memory_workbook_to_eggshell( in_mem_wb , newname , original_excel_name = None ):

  output_root = os.path.join( docs_root_get() , newname )

  if not os.path.exists( docs_root_get() ):
    os.system( 'mkdir ' + docs_root_get() )  #  e.g., mkdir /tmp/eggs/
  if not os.path.exists( output_root ):
    os.system( 'mkdir ' + output_root )  #  e.g., mkdir /tmp/eggs/dummy

  for (tab,cellname)  in  in_mem_wb :
    cc = in_mem_wb[(tab,cellname)]
    if tab == 'Belgium' and cellname == 'E13':
      print '\nwrite in memory workbook to eggshell : cc =',cc,'\n\n'
    cell_write( newname , tab , cellname , cc.value , species=cc.species , commit=False )

  powertable_workflow.origin_locate_and_record( output_root )
  if original_excel_name != None:
    powertable_workflow.excel_name_record( original_excel_name )
  versioning.initialize( output_root )

  return output_root

###################################################################################

def  available_loaded_docs():

  loaded_doc_paths = glob.glob( os.path.join( docs_root_get() , '*' ) )
  loaded_doc_paths = [ x.split('/')[-1] for x in loaded_doc_paths ]

  import projects

  if '' != projects.filter_get():
    loaded_doc_paths = filter( lambda d: (projects.filter_get().lower() in d.lower()) , loaded_doc_paths )

  return loaded_doc_paths
  
###################################################################################

def  available_sheets_in_doc( d , justrootsnotpaths = False ):

  sheets = glob.glob( os.path.join( docs_root_get() , d , '*' ) )
  if justrootsnotpaths:
    sheets = [ p.split('/')[-1] for p in sheets ]
  return sheets

###################################################################################

def  available_cells_in_egg_sheet ( d , s , justfilesnotpaths = False ):

  cells = glob.glob( os.path.join( docs_root_get() , d , s , '*' ) )
  if justfilesnotpaths:
    cells = [ p.split('/')[-1] for p in cells ]
  return cells

###################################################################################

def  filepath_for_egg ( doc ):
  return os.path.join( docs_root_get() , doc )

###################################################################################

def  make_eggshell_from_excel_file( path , filename_eggshell ):

  #  path is in the local filesystem, where an Excel workbook (xlsx file) can be found.
  #  filename_eggshell is what the eggshell should be named.
  print 'make eggshell from excel file :',path,filename_eggshell

  top    = docs_root_get()
  book   = openpyxl.load_workbook( path )
  path   = path.split('/')
  if len(path) == 0:
    print 'ERROR: path inadquate in make eggshell from excel file.'
  filename_excel = path[-1]
  eggroot = os.path.join( top , filename_eggshell )

  print 'path           =',path
  print 'filename_excel =',filename_excel
  print 'filename_eggsh =',filename_eggshell
  print 'eggroot        =',eggroot

  if not os.path.exists( top ):
    os.system( 'mkdir ' + top     )  #  e.g., mkdir /tmp/eggs/

  if not os.path.exists( eggroot ):
    os.system( 'mkdir ' + eggroot )  #  e.g., mkdir /tmp/eggs/dummy

  os.system( 'echo ' + filename_excel + ' > ' + os.path.join( eggroot, '.excel_filename' ) )

  for sheetname in book.sheetnames :

    sp = re.sub(' ','', sheetname)
    sp = sp.split('\.')[0]
    sp = eggroot + '/' + sp

    sheet = book.get_sheet_by_name( sheetname )
    dims  = [ sheet.max_column , sheet.max_row ]
    print 'sheetpath      =',sp,'of [#cols,#rows] =',dims

    if not os.path.exists( sp ):
      os.system( 'mkdir ' + sp + '/' )

    for   c in range(sheet.max_column):
      for r in range(sheet.max_row):

        cellname = coords_to_excel_cell_name( c , r )

        try:
          value = sheet[ cellname ].value
        except:
          print 'OOPS: problem reading cell value:',filename_eggshell,sheetname,cellname
          continue

        # http://stackoverflow.com/questions/23814522/cell-colors-using-openpyxl-2-02
        # http://openpyxl.readthedocs.org/en/latest/styles.html

        if value != None:

          cell_write( filename_eggshell , sheetname , cellname , value , species = None , commit = False )

  powertable_workflow.origin_locate_and_record( eggroot )
  versioning.initialize( eggroot )

###################################################################################

def  receive_file_from_browser( egg_name         ,
                                field            ,  #  i.e., role in project.
                                excel_file_name  ,
                                excel_file_bytes ,
                                project_name      
                              ):
# os.system( '[ -e ' + excel_folder + ' ] || mkdir ' + excel_folder )
  print 'RFFB : egg_name              :' , egg_name
  print 'RFFB : field                 :' , field
  print 'RFFB : excel_file_name       :' , excel_file_name
  print 'RFFB : len(excel_file_bytes) :' , len(excel_file_bytes)
  print 'RFFB : project_name          :' , project_name
  upload_to_path = os.path.join( '/tmp' , excel_file_name )
  print 'RFFB : upload_to_path        :' , upload_to_path
  try:
    with open( upload_to_path , 'w' ) as f:
      f.write( excel_file_bytes )
      f.flush()
  except:
    print '\n# error in do upload'
    print '', ; traceback.print_exc()

  if 'uploadedfilename_' in field and 'singleton' not in field:
    egg_name = re.sub( 'uploadedfilename_' , '' , field )
  make_eggshell_from_excel_file( upload_to_path , project_name + '__' + egg_name )

###################################################################################

def  do_clone_at_version( args ):
  #  'selected_eggshell' , 'selected_version' , 'masterfilename' are in args for sure.
  pass
  print '# IMPLEMENT  Tool_for_Excel.do_clone_at_version(',sanitized(args),')'

###################################################################################

def  filepath_for_doc( doc ):

  # doc may have been given as a full path on a commandline:
  # reduce it to just the filename, so we can use doc root.
  #
  # I may revisit this decision of 2015 Jan 26 soon.

  if '/' in doc:
    doc = doc.split('/')
    if doc[-1] == '':  #  e.g., '/tmp/eggs/blah/'  ->  [ 'tmp' , 'eggs' , 'blah' , '' ] .
      doc = doc[:-1]   #  i.e.,                        [ 'tmp' , 'eggs' , 'blah'      ] .
    doc = doc[-1]      #  i.e.,                                           'blah'        .

  path = os.path.join( docs_root_get() , doc )
  path = re.sub('\/\/','/',path)  #  clean up any '//' to single '/'.
  return path

###################################################################################

def  filepath_for_sheet( doc , sheet ):

  return os.path.join( filepath_for_doc( doc ) , sheet )

###################################################################################
 
def  filepath_for_cell( doc , sheet , cell ):  #  this works as is for 'fs' and 'dbm' representations.

  return os.path.join( filepath_for_sheet( doc , sheet ) , cell )

###################################################################################

def  species_from_eggname( egg ):
  result = (egg.split('__')[-1]).lower()
# print 'DEBUG species_from_eggname(',egg,') =',result
  return result

###################################################################################

class Cell:
  def  __init__( this , v , s ):
    this.value   = v
    this.species = s
  def  __repr__( this ):
    return 'Cell(' + str( this.value ) + ',' + str(this.species) + ')'
  def  style   ( this ): # return a css style.  Odd since this file is normally html unaware.
    import powertable_workflow 
    return powertable_workflow.cssstyleforsource( this.species )

###################################################################################

def  cell_read  ( doc , sheet , cell , whiney=False ):

    if '/' in doc:
      doc = doc.split('/')[-1]

    cell_path = filepath_for_cell( doc , sheet , cell )

    try:
      with open(cell_path,'r') as f:
        lines  = [ v.strip() for v in f.read().strip().split('\n') ]
        if sheet == 'Belgium' and cell == 'E13' :
          print 'CELL READ',doc,sheet,cell,': lines =',lines
        source = None
        if len(lines) > 1:
          source = lines[1]
        value = Cell( lines[0] , source )
        if whiney:
          print 'JUST READ value \'',value.value,'\' (',type(value.value),') from',doc,sheet,cell
    except:
      if whiney:
        print '# do read ERROR (fs)  :  %s did not exist.' % cell_path
      value = Cell('',None)

    return value 

###################################################################################

def  cell_write ( doc , sheet , cellname , value , species = None ,
                  commit=True , as_user='noone' , debug=False ):

    # if doc has precursors, ignore edits, because it is a master/derived document.
    doc = doc.split('/')[-1]  #  just the egg name, not its path.

#   if debug and len( file_dependence.precursors_of( doc ) ) > 0:
#     print '# IGNORING edit to derived document',doc
#     return

    doc_path = filepath_for_doc( doc )
    if not os.path.isdir( doc_path ):
      os.system( 'mkdir ' + doc_path )

    sheet_path = os.path.join( doc_path , sheet )
    if not os.path.isdir( sheet_path ):
      os.system( 'mkdir ' + sheet_path )

    cell_path  = os.path.join( sheet_path , cellname )

    try:
      with open(cell_path,'w') as f:
        bytes = str(value)
        if species:
          bytes = bytes + '\n' + str(species)
        f.write(bytes)
        try:
          log( doc , { cellname : (value,species) } )
        except:
          print '# what the heck is with the log in cell write ?', cell_path , value
    except:
      print '# do write ERROR (fs)  :  %s did not exist.' % cell_path

    if commit:
      versioning.add_commit( doc , as_user )

    powertable_workflow.propagate_cell_update_from( doc , sheet , cellname )

###################################################################################

def  do_writes( vars  , as_user='noone' ):

    for lhs in vars:
      if 'c_' == lhs[:2] :
        print '# Tools for Excel cell write :  ',vars['doc'],vars['tab'],lhs[2:],vars[lhs]
        cell_write( vars['doc'] , vars['tab'] , lhs[2:] , vars[lhs] , commit=False)

    if 'doc' in vars:
      log( vars['doc'] , vars )
      versioning.add_commit( vars['doc'] , as_user )

###################################################################################

def  do_list():
  docs            = available_loaded_docs()
  docs_and_sheets = []

  for d in docs:
    s = available_sheets_in_doc(d)
    if 'verbose' not in sys.argv:
      s = [ p.split('/')[-1] for p in s ]
    docs_and_sheets.append( [ d , s ] )

  return docs_and_sheets

###################################################################################

def  do_excelify( doc , path ):  #  OPENPYXL

  sheets = available_sheets_in_doc( doc )
 #print '# sheets available in ', doc, ' are ', sheets

  from xlwt import Workbook
  book = xlwt.Workbook()

  for s in sheets:
    s = s.split('/')[-1]
    shit = book.add_sheet(s)

    cells = available_cells_in_egg_sheet( doc , s )
    print '# %s cells:' % s
    t     = [ c.split('/')[-1] for c in cells ]
    t     = [ ( c , excel_cell_name_to_coords( c ) , cell_read( doc , s , c ).value ) for c in t ]
   #print t
    for c in t:
      v = c[2]

      try:
        v = float(v)
        failed = False
      except:
        v = c[2]
        failed = True

      if failed:
        try:
          v = int(v)
          failed = False
        except:
          v = c[2]
          failed = True

      shit.write( c[1][1], c[1][0] , v )

  if '.xls' not in path:
    path = path + '.xls'

  book.save(path)
  print '# saved %s (or at least i think i did) from the EggShell doc %s.' % (path,doc)

###################################################################################

def  do_erase ( doc , username ):

    docs = available_loaded_docs()
    print '\n# docs_root =',docs_root_get()
    print '# doc       =',doc
    print '# docs loaded and available (pre erase) are ', docs
    if doc in docs:
      dr = os.path.join( docs_root_get() , doc )
      print '# /bin/rm -fr ' + dr
      os.system( '/bin/rm -fr ' + dr )
    file_dependence.remove_file( doc )

###################################################################################

def  do_clean_generated_download_files():
  files = glob.glob( '/tmp/*.xls' )
  for f in files:
    os.system( '/bin/rm ' + f )

###################################################################################

def  log( doc , dict ):

  # Only log dict if it holds a lhs that starts with 'c_'.

  username = 'nobody'
  if 'username' in dict:
    username = dict['username']

  worth_logging = False
  for lhs in dict:
    if 'c_' == lhs[:2] :
      worth_logging = True
      break

  if worth_logging:

    dict = copy.deepcopy(dict)
    dict['version'    ] = now_as_a_string()
    dict['server_euid'] = getpass.getuser()
    

    try:
      logfilename = os.path.join( docs_root_get() , doc , '.log' )
      with open( logfilename , 'a' ) as logfile:
        logfile.write( str(sanitized(dict)) + '\n' )
    except:
      print '# There is a bug in log.',doc,sanitized(dict)

###################################################################################

#  EggShell persistance layout in the filesystem:
#
#    /workbook_name/sheet_name/cell_name
#
#  where the cell_name is a text file
#  with the value of the cell as its single line.

###################################################################################

if  '__main__' == __name__ :

# print len(sys.argv), ' is len(sys.argv) '
# print '\nCONSIDER HDF5 OR WHATEVER IT IS NAMED FOR THE FILE FORMAT\n'

  if   (1 == len(sys.argv)) or ('help' in sys.argv):
    n = sys.argv[0]
    print '''
     %s is jozwiak@apple.com's spring 2015 command line tool for Microsoft Excel.

     Usage is as follows.

     %s  help                                             #   prints this message.
     %s  describe_excel  d                                #   give dimensions for each sheet in d:  if d is a path, d must be an Excel document; otherwise it is assumed already loaded for the web.
     %s  load     path                                    #   reads in  the Excel document at path.
     %s  export   doc  path                               #   make an Excel document from eggshell document doc;  write the Excel file at path.
     %s  list [verbose]                                   #   lists the names of the loaded Excel documents, and the sheets in each document.  If verbose is given, full paths to the sheets are given.
     %s  read  doc sheet cell                             #   doc is a loaded workbook's name, such as J86v2, sheet might be Korea, cell perhaps G45.
     %s  write doc sheet cell  value_string               #   doc is a loaded workbook's name, such as J86v2, sheet might be Korea, cell perhaps G45, value perhaps 136.96 .

     %s  log       doc                                    #   shows the history of edits of doc, with associated version numbers.
     %s  flashback doc hexstring                          #   shows the doc as of point-in-time hexstring.
     %s  now       doc                                    #   shows the doc at its latest.

     %s  combine folder                                   #   expects 4 Excel documents of matching dimensions in folder, and generates 1 with name folder_combined, for WLAN legacy nonsense.

     %s  diff    doc  hexstring                           #   shows the tabs and cells in each tab changed over a time span.
     %s  diffs   doc  hexstring_e  hexstring_l            #   shows the tabs and cells in each tab changed over a time span.

     %s  insert [row|column] [before|after] integer in d  #   NOT AVAILABLE YET:  will do the obvious.
     %s  delete [row|column]                integer in d  #   NOT AVAILABLE YET:  will do the obvious.
     ''' % (n,n,n,n,n,n,n,n,n,n,n,n,n,n,n,n)
    sys.exit(0)

  elif 'describe_excel' in sys.argv:
    #
    #   describe eggshelldoc
    #
    if len(sys.argv) < 3:
      print '# Error:  an Eggshell doc name is needed.  Fail.'
      sys.exit(1)
    else:
      d = sys.argv[2]
      describe_excel( d )

  elif 'load' in sys.argv:
    #
    #   load eggshelldoc
    #
    if len(sys.argv) < 3:
      print '# Error:  an Excel, not Eggshell, doc name is needed.  Fail.'
      sys.exit(1)
    else:
      egg = sys.argv[2].split('/')[-1].split('.')[0]
      make_eggshell_from_excel_file( sys.argv[2] , egg )

  elif 'export' in sys.argv:
    #
    #   export eggshelldoc path
    #
    if len(sys.argv) < 4:
      print '# Error:  an Eggshell doc name is needed as is a filesystem path to name the output file.  Fail.'
      docs = glob.glob( docs_root_get() + '*' )
      if len(docs) > 0:
        print '#         Here is a list of the loaded docs.'
        print '#'
        for d in docs:
          print '#           ', d , ' ( ' , d.split('/')[-1] , ' ) '
        print '#'
        print
      else:
        print '#         There are no loaded docs, by the way.'
      sys.exit(1)
    else:
      do_excelify( sys.argv[2] , sys.argv[3] )

  elif 'list' in sys.argv:
    #
    #   list eggshelldoc
    #
    print json.dumps( do_list() , indent=3 )

  elif 'read'   in sys.argv:
     #
     #   read  path sheet cell
     #
    if len(sys.argv) < 5:
      print '# Error.  Try %s read doc sheet cell.' % sys.argv[0]
      print '# possible documents are',
      for d in do_list():
        print d[0],
      print
      sys.exit(1)
    else:
      value = cell_read ( sys.argv[2] , sys.argv[3] , sys.argv[4] ).value
      print value
      #                 doc         , sheet       , cell

  elif 'write'  in sys.argv:
     #
     #   write path sheet cell value_string
     #
    if len(sys.argv) < 6:
      print '# Error.  Try %s write doc sheet cell value_string.' % sys.argv[0]
      print '# possible documents are',
      for d in do_list():
        print d[0],
      print
      sys.exit(1)
    else:
      cell_write( sys.argv[2] , sys.argv[3] , sys.argv[4] , sys.argv[5] ) 
      #         doc         , sheet       , cell        , value

  elif 'insert' in sys.argv:
    #
    #   insert [row|column] [before|after] integer in eggshelldoc
    #
    if len(sys.argv) < 7:
      print '# Error:  an Eggshell doc name is needed.  Fail.'
      sys.exit(1)
    else:
      pass

  elif 'delete' in sys.argv:
    #
    #   delete [row|column] integer in eggshelldoc
    #
    if len(sys.argv) < 6:
      print '# Error:  an Eggshell doc name is needed.  Fail.'
      sys.exit(1)
    else:
      pass 

  elif 'log' in sys.argv:
    if len(sys.argv) < 3:
      print '# Error:  an Eggshell doc name is needed.  Fail.'
      sys.exit(1)
    else:
      t = versioning.do_history ( sys.argv[2] )
      for line in t:
        print '# ',line

  elif 'flashback' in sys.argv:
    if len(sys.argv) < 4:
      print '# Error:  a doc name and a hexstring specifying a point in time in the history of doc are needed.  Fail.'
      sys.exit(1)
    else:
      versioning.do_flashback( sys.argv[2] , sys.argv[3] )

  elif 'now' in sys.argv:
    if len(sys.argv) < 3:
      print '# Error:  an Eggshell doc name is needed.  Fail.'
      sys.exit(1)
    else:
      versioning.do_back_to_now ( sys.argv[2] )

  elif 'diff' in sys.argv:
    if len(sys.argv) < 4:
      print '# Error:  an Eggshell doc name is needed, as is a version of that doc.  Fail.'
      sys.exit(1)
    else:
      print ( versioning.do_diffs_latest_vs_older( sys.argv[2] , sys.argv[3] ) ).structure()

  elif 'difff' in sys.argv:
    if len(sys.argv) < 5:
      print '# Error:  an Eggshell doc name is needed, as are two versions of that doc.  Fail.'
      sys.exit(1)
    else:
      print ( versioning.paths_changed_over_timespan ( sys.argv[2] , sys.argv[3] , sys.argv[4] ) ).structure()

else:
  print '# imported tools for Eggshell...' 

###################################################################################
